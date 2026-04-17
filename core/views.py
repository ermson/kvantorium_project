from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count
from django.db.models.functions import ExtractMonth
from django.http import HttpResponse
from .models import Student, Project, ProjectStage, Mentor
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
import os
from django.conf import settings



# ==================== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ====================

def get_urgency_status(planned_date):
    if not planned_date:
        return 'secondary'
    today = date.today()
    days_left = (planned_date - today).days
    if days_left < 0:
        return 'danger'
    elif days_left <= 30:
        return 'warning'
    else:
        return 'secondary'


def get_user_mentor(request):
    if not request.user.is_authenticated:
        return None
    try:
        return request.user.mentor
    except:
        return None


def user_can_edit(request):
    """Проверяет, может ли пользователь редактировать данные"""
    # Суперпользователь может всё
    if request.user.is_superuser:
        return True
    # Пользователь с логином metodist НЕ может редактировать
    if request.user.username == 'metodist':
        return False
    # Все остальные могут редактировать
    return True


# ==================== АВТОРИЗАЦИЯ ====================

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            return render(request, 'core/login.html', {'error': 'Неверный логин или пароль'})
    return render(request, 'core/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


# ==================== ДАШБОРД ====================

@login_required
def dashboard(request):
    current_mentor = get_user_mentor(request)
    is_superuser = request.user.is_superuser
    is_methodist = request.user.groups.filter(name='Методист').exists()
    can_edit = user_can_edit(request)

    if is_methodist:
        students_count = Student.objects.count()
        projects = Project.objects.all()
        projects_count = projects.count()
        in_progress_count = projects.filter(status='development').count()
        urgent_stages = ProjectStage.objects.filter(is_completed=False).count()
        urgent_projects = projects.filter(status__in=['idea', 'development'])[:5]
        status_counts = projects.values('status').annotate(count=Count('id'))
        monthly_stages = ProjectStage.objects.filter(
            planned_date__isnull=False
        ).annotate(month=ExtractMonth('planned_date')).values('month').annotate(count=Count('id')).order_by('month')
        today = date.today()
        urgent_stages_list = ProjectStage.objects.filter(
            is_completed=False,
            planned_date__gte=today,
            planned_date__lte=today + timedelta(days=30)
        ).order_by('planned_date')[:10]
    else:
        if not is_superuser and current_mentor:
            students_count = Student.objects.filter(mentor=current_mentor).count()
            projects = Project.objects.filter(student__mentor=current_mentor)
            projects_count = projects.count()
            in_progress_count = projects.filter(status='development').count()
            urgent_stages = ProjectStage.objects.filter(project__student__mentor=current_mentor,
                                                        is_completed=False).count()
            urgent_projects = projects.filter(status__in=['idea', 'development'])[:5]
            status_counts = projects.values('status').annotate(count=Count('id'))
            monthly_stages = ProjectStage.objects.filter(
                project__student__mentor=current_mentor,
                planned_date__isnull=False
            ).annotate(month=ExtractMonth('planned_date')).values('month').annotate(count=Count('id')).order_by('month')
            today = date.today()
            urgent_stages_list = ProjectStage.objects.filter(
                project__student__mentor=current_mentor,
                is_completed=False,
                planned_date__gte=today,
                planned_date__lte=today + timedelta(days=30)
            ).order_by('planned_date')[:10]
        else:
            students_count = Student.objects.count()
            projects = Project.objects.all()
            projects_count = projects.count()
            in_progress_count = projects.filter(status='development').count()
            urgent_stages = ProjectStage.objects.filter(is_completed=False).count()
            urgent_projects = projects.filter(status__in=['idea', 'development'])[:5]
            status_counts = projects.values('status').annotate(count=Count('id'))
            monthly_stages = ProjectStage.objects.filter(
                planned_date__isnull=False
            ).annotate(month=ExtractMonth('planned_date')).values('month').annotate(count=Count('id')).order_by('month')
            today = date.today()
            urgent_stages_list = ProjectStage.objects.filter(
                is_completed=False,
                planned_date__gte=today,
                planned_date__lte=today + timedelta(days=30)
            ).order_by('planned_date')[:10]

    status_list = []
    for item in status_counts:
        status_display = dict(Project.STATUS_CHOICES).get(item['status'], item['status'])
        status_list.append((status_display, item['count']))

    months_dict = {1: 'Янв', 2: 'Фев', 3: 'Мар', 4: 'Апр', 5: 'Май', 6: 'Июн', 7: 'Июл', 8: 'Авг', 9: 'Сен', 10: 'Окт',
                   11: 'Ноя', 12: 'Дек'}
    months_list = []
    for item in monthly_stages:
        month_num = item['month']
        month_name = months_dict.get(month_num, str(month_num))
        months_list.append((month_name, item['count']))

    context = {
        'students_count': students_count,
        'projects_count': projects_count,
        'in_progress_count': in_progress_count,
        'urgent_stages': urgent_stages,
        'urgent_projects': urgent_projects,
        'urgent_stages_list': urgent_stages_list,
        'is_superuser': is_superuser,
        'current_mentor': current_mentor,
        'can_edit': can_edit,
        'now': date.today(),
        'status_counts': status_list,
        'monthly_stages': months_list,
    }
    return render(request, 'core/dashboard.html', context)


# ==================== УВЕДОМЛЕНИЯ ====================

def notifications_context(request):
    if not request.user.is_authenticated:
        return {'notifications': [], 'urgent_notifications_count': 0, 'can_edit': False}

    today = date.today()
    one_month_later = today + timedelta(days=30)
    current_mentor = get_user_mentor(request)
    is_superuser = request.user.is_superuser
    is_methodist = request.user.groups.filter(name='Методист').exists()
    can_edit = user_can_edit(request)

    if is_methodist:
        urgent_stages = ProjectStage.objects.filter(
            is_completed=False,
            planned_date__gte=today,
            planned_date__lte=one_month_later
        ).select_related('project', 'project__student').order_by('planned_date')
    elif not is_superuser and current_mentor:
        students_ids = Student.objects.filter(mentor=current_mentor).values_list('id', flat=True)
        projects_ids = Project.objects.filter(student_id__in=students_ids).values_list('id', flat=True)
        urgent_stages = ProjectStage.objects.filter(
            is_completed=False,
            project_id__in=projects_ids,
            planned_date__gte=today,
            planned_date__lte=one_month_later
        ).select_related('project', 'project__student').order_by('planned_date')
    else:
        urgent_stages = ProjectStage.objects.filter(
            is_completed=False,
            planned_date__gte=today,
            planned_date__lte=one_month_later
        ).select_related('project', 'project__student').order_by('planned_date')

    notifications = []
    for stage in urgent_stages:
        days_left = (stage.planned_date - today).days
        notifications.append({
            'id': stage.id,
            'message': f"⚠️ Через {days_left} дн. этап «{stage.stage_name}» проекта «{stage.project.name}»",
            'url': f"/stages/{stage.id}/",
            'days_left': days_left,
        })

    return {
        'notifications': notifications[:15],
        'urgent_notifications_count': len(notifications),
        'can_edit': can_edit,
    }


# ==================== УЧЕНИКИ ====================

@login_required
def student_list(request):
    students = Student.objects.all()

    current_mentor = get_user_mentor(request)
    is_superuser = request.user.is_superuser
    is_methodist = request.user.groups.filter(name='Методист').exists()
    can_edit = user_can_edit(request)
    mentor_filter = request.GET.get('mentor_filter', '')

    if is_methodist or mentor_filter == 'all':
        students = Student.objects.all()
    elif mentor_filter and mentor_filter != 'my':
        try:
            students = students.filter(mentor_id=int(mentor_filter))
        except:
            pass
    else:
        if not is_superuser and current_mentor:
            students = students.filter(mentor=current_mentor)

    search_query = request.GET.get('search', '')
    if search_query:
        students = students.filter(
            Q(last_name__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(middle_name__icontains=search_query)
        )

    group_filter = request.GET.get('group', '')
    if group_filter:
        students = students.filter(group_name=group_filter)

    groups = Student.objects.values_list('group_name', flat=True).distinct()
    mentors = Mentor.objects.all()

    context = {
        'students': students,
        'search_query': search_query,
        'group_filter': group_filter,
        'groups': groups,
        'mentors': mentors,
        'mentor_filter': mentor_filter,
        'current_mentor_id': current_mentor.id if current_mentor else None,
        'is_superuser': is_superuser,
        'can_edit': can_edit,
        'is_methodist': is_methodist,
    }
    return render(request, 'core/student_list.html', context)


@login_required
def student_detail(request, pk):
    student = get_object_or_404(Student, pk=pk)
    projects = student.project_set.all()
    can_edit = user_can_edit(request)
    return render(request, 'core/student_detail.html', {'student': student, 'projects': projects, 'can_edit': can_edit})


@login_required
def student_create(request):
    if not user_can_edit(request):
        return redirect('student_list')

    if request.method == 'POST':
        student = Student.objects.create(
            last_name=request.POST['last_name'],
            first_name=request.POST['first_name'],
            middle_name=request.POST.get('middle_name', ''),
            group_name=request.POST['group_name'],
            mentor_id=request.POST['mentor'],
            enrollment_date=request.POST['enrollment_date']
        )
        return redirect('student_list')

    mentors = Mentor.objects.all()
    group_choices = [
        'Промышленная робототехника',
        'VR/AR (Виртуальная реальность)',
        'Хайтек (High-tech)',
        'IT (Информационные технологии)',
        'Био (Биотехнологии)',
        'Промдизайн (Промышленный дизайн)',
    ]
    return render(request, 'core/student_form.html', {'mentors': mentors, 'group_choices': group_choices})


@login_required
def student_edit(request, pk):
    if not user_can_edit(request):
        return redirect('student_list')

    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.last_name = request.POST['last_name']
        student.first_name = request.POST['first_name']
        student.middle_name = request.POST.get('middle_name', '')
        student.group_name = request.POST['group_name']
        student.mentor_id = request.POST['mentor']
        student.enrollment_date = request.POST['enrollment_date']
        student.save()
        return redirect('student_list')

    mentors = Mentor.objects.all()
    group_choices = [
        'Промышленная робототехника',
        'VR/AR (Виртуальная реальность)',
        'Хайтек (High-tech)',
        'IT (Информационные технологии)',
        'Био (Биотехнологии)',
        'Промдизайн (Промышленный дизайн)',
    ]
    return render(request, 'core/student_form.html',
                  {'student': student, 'mentors': mentors, 'group_choices': group_choices})


@login_required
def student_delete(request, pk):
    if not user_can_edit(request):
        return redirect('student_list')

    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.delete()
        return redirect('student_list')
    return render(request, 'core/confirm_delete.html', {'object': student, 'type': 'student'})


# ==================== ПРОЕКТЫ ====================

@login_required
def project_list(request):
    """Список проектов с фильтром по наставнику"""
    projects = Project.objects.all()

    current_mentor = get_user_mentor(request)
    is_superuser = request.user.is_superuser
    is_methodist = request.user.groups.filter(name='Методист').exists()
    can_edit = user_can_edit(request)

    # Получаем параметр фильтра из GET-запроса
    mentor_filter = request.GET.get('mentor_filter', '')

    # Логика фильтрации
    if is_methodist or mentor_filter == 'all':
        projects = Project.objects.all()
    elif mentor_filter and mentor_filter != 'my':
        try:
            projects = projects.filter(student__mentor_id=int(mentor_filter))
        except:
            pass
    else:
        if not is_superuser and current_mentor:
            projects = projects.filter(student__mentor=current_mentor)

    # Поиск по названию
    search_query = request.GET.get('search', '')
    if search_query:
        projects = projects.filter(name__icontains=search_query)

    # Фильтр по статусу
    status_filter = request.GET.get('status', '')
    if status_filter:
        projects = projects.filter(status=status_filter)

    status_choices = Project.STATUS_CHOICES
    mentors = Mentor.objects.all()

    context = {
        'projects': projects,
        'search_query': search_query,
        'status_filter': status_filter,
        'status_choices': status_choices,
        'mentors': mentors,
        'mentor_filter': mentor_filter,
        'is_superuser': is_superuser,
        'can_edit': can_edit,
        'is_methodist': is_methodist,
    }
    return render(request, 'core/project_list.html', context)


@login_required
def project_detail(request, pk):
    project = get_object_or_404(Project, pk=pk)
    stages = project.projectstage_set.all().order_by('planned_date')
    today = date.today()
    can_edit = user_can_edit(request)

    for stage in stages:
        if stage.is_completed:
            stage.urgency_status = 'success'
            stage.days_left = None
        else:
            days_left = (stage.planned_date - today).days
            stage.days_left = days_left
            if days_left < 0:
                stage.urgency_status = 'danger'
            elif days_left <= 7:
                stage.urgency_status = 'danger'
            elif days_left <= 30:
                stage.urgency_status = 'warning'
            else:
                stage.urgency_status = 'secondary'

    context = {
        'project': project,
        'stages': stages,
        'can_edit': can_edit,
    }
    return render(request, 'core/project_detail.html', context)


@login_required
def project_create(request):
    if not user_can_edit(request):
        return redirect('project_list')

    if request.method == 'POST':
        project = Project.objects.create(
            name=request.POST['name'],
            description=request.POST['description'],
            student_id=request.POST['student'],
            status=request.POST['status']
        )
        return redirect('project_list')
    students = Student.objects.all()
    return render(request, 'core/project_form.html', {'students': students})


@login_required
def project_edit(request, pk):
    if not user_can_edit(request):
        return redirect('project_list')

    project = get_object_or_404(Project, pk=pk)
    if request.method == 'POST':
        project.name = request.POST['name']
        project.description = request.POST['description']
        project.student_id = request.POST['student']
        project.status = request.POST['status']
        project.save()
        return redirect('project_list')
    students = Student.objects.all()
    return render(request, 'core/project_form.html', {'project': project, 'students': students})


@login_required
def project_delete(request, pk):
    if not user_can_edit(request):
        return redirect('project_list')

    project = get_object_or_404(Project, pk=pk)
    if request.method == 'POST':
        project.delete()
        return redirect('project_list')
    return render(request, 'core/confirm_delete.html', {'object': project, 'type': 'project'})


# ==================== ЭКСПОРТ ====================

@login_required
def export_projects_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Проекты"

    headers = ['ID', 'Название проекта', 'Ученик', 'Статус', 'Дата создания']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    projects = Project.objects.all().select_related('student')
    for row, project in enumerate(projects, 2):
        ws.cell(row=row, column=1, value=project.id)
        ws.cell(row=row, column=2, value=project.name)
        ws.cell(row=row, column=3, value=f"{project.student.last_name} {project.student.first_name}")
        ws.cell(row=row, column=4, value=project.get_status_display())
        ws.cell(row=row, column=5, value=project.created_at.strftime('%d.%m.%Y'))

    for col in range(1, 6):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="projects_export.xlsx"'
    wb.save(response)
    return response

@login_required
def export_project_pdf(request, pk):
    """Экспорт отчёта по проекту в PDF на русском языке"""
    project = get_object_or_404(Project, pk=pk)
    stages = project.projectstage_set.all().order_by('planned_date')

    # Регистрируем русский шрифт
    font_path = os.path.join(settings.BASE_DIR, 'fonts', 'arial.ttf')
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont('Arial', font_path))
        font_name = 'Arial'
    else:
        font_name = 'Helvetica'

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="project_{project.id}_report.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4,
                            topMargin=1.5 * cm, bottomMargin=1.5 * cm,
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    styles = getSampleStyleSheet()

    # Стили с русским шрифтом
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontName=font_name, fontSize=14, alignment=1,
                                 spaceAfter=15)
    heading_style = ParagraphStyle('HeadingStyle', parent=styles['Heading3'], fontName=font_name, fontSize=11,
                                   spaceAfter=8, textColor=colors.HexColor('#2c3e50'))
    normal_style = ParagraphStyle('NormalStyle', parent=styles['Normal'], fontName=font_name, fontSize=9, spaceAfter=4)

    elements = []

    # Заголовок
    elements.append(Paragraph(f"Отчёт по проекту: {project.name}", title_style))
    elements.append(Spacer(1, 5))

    # 1. Информация о проекте
    elements.append(Paragraph("1. Информация о проекте", heading_style))
    elements.append(Spacer(1, 3))

    project_data = [
        ["Название:", project.name],
        ["Описание:",
         (project.description or "Нет описания")[:150] + "..." if len(project.description or "") > 150 else (
                     project.description or "Нет описания")],
        ["Статус:", project.get_status_display()],
        ["Дата создания:", project.created_at.strftime('%d.%m.%Y')],
        ["Последнее обновление:", project.updated_at.strftime('%d.%m.%Y')],
    ]

    project_table = Table(project_data, colWidths=[3.5 * cm, 11 * cm])
    project_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(project_table)
    elements.append(Spacer(1, 12))

    # 2. Ученик и наставник
    elements.append(Paragraph("2. Ученик и наставник", heading_style))
    elements.append(Spacer(1, 3))

    student_data = [
        ["Ученик:", f"{project.student.last_name} {project.student.first_name}"],
        ["Направление:", project.student.group_name],
        ["Дата поступления:", project.student.enrollment_date.strftime('%d.%m.%Y')],
        ["Наставник:",
         f"{project.student.mentor.user.last_name} {project.student.mentor.user.first_name}" if project.student.mentor else "Не назначен"],
        ["Квантум:", project.student.mentor.department if project.student.mentor else "—"],
        ["Телефон:", project.student.mentor.phone if project.student.mentor else "—"],
    ]

    student_table = Table(student_data, colWidths=[3.5 * cm, 11 * cm])
    student_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(student_table)
    elements.append(Spacer(1, 12))

    # 3. Этапы проекта
    elements.append(Paragraph("3. Этапы проекта", heading_style))
    elements.append(Spacer(1, 3))

    # Заголовки таблицы этапов
    stage_table_data = [["№", "Название этапа", "Плановая дата", "Фактическая дата", "Статус"]]
    for i, stage in enumerate(stages, 1):
        status = "Выполнен" if stage.is_completed else "В работе"
        stage_table_data.append([
            str(i),
            stage.stage_name[:40] + "..." if len(stage.stage_name) > 40 else stage.stage_name,
            stage.planned_date.strftime('%d.%m.%y'),
            stage.actual_date.strftime('%d.%m.%y') if stage.actual_date else "—",
            status,
        ])

    stage_table = Table(stage_table_data, colWidths=[0.8 * cm, 6.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm])
    stage_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(stage_table)
    elements.append(Spacer(1, 12))

    # 4. Прогресс выполнения
    completed = stages.filter(is_completed=True).count()
    total = stages.count()
    progress = int((completed / total) * 100) if total > 0 else 0

    elements.append(Paragraph("4. Прогресс выполнения", heading_style))
    elements.append(Spacer(1, 3))
    elements.append(Paragraph(f"Выполнено этапов: {completed} из {total} ({progress}%)", normal_style))
    elements.append(Spacer(1, 15))
    elements.append(Paragraph(f"Отчёт сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}", normal_style))

    doc.build(elements)
    return response
# ==================== ЭТАПЫ ====================

@login_required
def stage_list(request):
    """Список этапов с точным количеством дней до дедлайна"""
    stages = ProjectStage.objects.all().order_by('planned_date')
    today = date.today()

    current_mentor = get_user_mentor(request)
    is_superuser = request.user.is_superuser
    is_methodist = request.user.groups.filter(name='Методист').exists()
    can_edit = user_can_edit(request)

    # Получаем параметр фильтра из GET-запроса
    mentor_filter = request.GET.get('mentor_filter', '')

    # Логика фильтрации
    if is_methodist or mentor_filter == 'all':
        stages = ProjectStage.objects.all().order_by('planned_date')
    elif mentor_filter and mentor_filter != 'my':
        try:
            stages = stages.filter(project__student__mentor_id=int(mentor_filter)).order_by('planned_date')
        except:
            pass
    else:
        if not is_superuser and current_mentor:
            stages = stages.filter(project__student__mentor=current_mentor).order_by('planned_date')

    # Поиск по названию этапа
    search_query = request.GET.get('search', '')
    if search_query:
        stages = stages.filter(stage_name__icontains=search_query)

    # Фильтр по выполнению
    completed_filter = request.GET.get('completed', '')
    if completed_filter == 'completed':
        stages = stages.filter(is_completed=True)
    elif completed_filter == 'not_completed':
        stages = stages.filter(is_completed=False)

    # Добавляем статус срочности и количество дней
    for stage in stages:
        if stage.is_completed:
            stage.urgency_status = 'success'
            stage.days_left = None
        else:
            days_left = (stage.planned_date - today).days
            stage.days_left = days_left
            if days_left < 0:
                stage.urgency_status = 'danger'
            elif days_left <= 7:
                stage.urgency_status = 'danger'
            elif days_left <= 30:
                stage.urgency_status = 'warning'
            else:
                stage.urgency_status = 'secondary'

    mentors = Mentor.objects.all()

    context = {
        'stages': stages,
        'search_query': search_query,
        'completed_filter': completed_filter,
        'today': today,
        'mentors': mentors,
        'mentor_filter': mentor_filter,
        'is_superuser': is_superuser,
        'can_edit': can_edit,
        'is_methodist': is_methodist,
    }
    return render(request, 'core/stage_list.html', context)


@login_required
def stage_detail(request, pk):
    stage = get_object_or_404(ProjectStage, pk=pk)
    can_edit = user_can_edit(request)
    return render(request, 'core/stage_detail.html', {'stage': stage, 'can_edit': can_edit})


@login_required
def stage_create(request):
    if not user_can_edit(request):
        return redirect('stage_list')

    if request.method == 'POST':
        stage = ProjectStage.objects.create(
            project_id=request.POST['project'],
            stage_name=request.POST['stage_name'],
            planned_date=request.POST['planned_date'],
            actual_date=request.POST.get('actual_date') or None,
            is_completed=request.POST.get('is_completed') == 'on',
            comment=request.POST.get('comment', '')
        )
        return redirect('stage_list')
    projects = Project.objects.all()
    return render(request, 'core/stage_form.html', {'projects': projects})


@login_required
def stage_edit(request, pk):
    if not user_can_edit(request):
        return redirect('stage_list')

    stage = get_object_or_404(ProjectStage, pk=pk)
    if request.method == 'POST':
        stage.project_id = request.POST['project']
        stage.stage_name = request.POST['stage_name']
        stage.planned_date = request.POST['planned_date']
        stage.actual_date = request.POST.get('actual_date') or None
        stage.is_completed = request.POST.get('is_completed') == 'on'
        stage.comment = request.POST.get('comment', '')
        stage.save()
        return redirect('stage_list')
    projects = Project.objects.all()
    return render(request, 'core/stage_form.html', {'stage': stage, 'projects': projects})


@login_required
def stage_delete(request, pk):
    if not user_can_edit(request):
        return redirect('stage_list')

    stage = get_object_or_404(ProjectStage, pk=pk)
    if request.method == 'POST':
        stage.delete()
        return redirect('stage_list')
    return render(request, 'core/confirm_delete.html', {'object': stage, 'type': 'stage'})


@login_required
def stage_mark_complete(request, pk):
    if not user_can_edit(request):
        return redirect('stage_list')

    stage = get_object_or_404(ProjectStage, pk=pk)
    stage.is_completed = True
    stage.actual_date = date.today()
    stage.save()
    return redirect('project_detail', pk=stage.project.id)